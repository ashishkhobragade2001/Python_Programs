import pandas
def read_excel_data(file_path):
    data = pandas.read_excel(file_path) 
    return data
## calling 
file_path = "C:\\ashish.xlsx"
data = read_excel_data(file_path)
#print("data from excel is :\n",data)

import pandas as pd  # Import pandas

def insert_data(file_path, value):
    # Read the existing Excel file
    data = pd.read_excel(file_path)

    # Convert the new value into a DataFrame
    new_data = pd.DataFrame([value])  # Wrap value in a list

    # Concatenate the new data with the existing data
    data = pd.concat([data, new_data], ignore_index=True)

    # Write the updated data back to the Excel file
    data.to_excel(file_path, index=False)
    print("Data has been inserted.")



# New value to be added
value = {'name': 'rahul', 'age': 41}

# Call the function
insert_data(file_path, value)




from openpyxl import load_workbook
def read_data_from_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook["Sheet1"]      # change accourding to sheet name
    for row in sheet.iter_rows(values_only=True):
        print(row)
       
#read_data_from_excel(file_path)