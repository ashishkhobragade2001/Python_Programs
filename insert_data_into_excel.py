from openpyxl import load_workbook

def insert_data_into_excel(file_path, new_data):

    workbook = load_workbook(filename=file_path)
    
    # Get the active sheet
    sheet = workbook.active  
    #sheet = workbook["ashish"]  # Specify the sheet name

    # Find the next empty row to start adding data
    next_row = sheet.max_row + 1
    
    # Add new data to the Excel file
    for insert in new_data:
        sheet.cell(row=next_row, column=1).value = insert["Name"]
        sheet.cell(row=next_row, column=2).value = insert["Age"]
        sheet.cell(row=next_row, column=3).value = insert["City"]
        next_row += 1
    
    # Save the changes
    workbook.save(filename=file_path)
    print(f"Data added successfully to {file_path}")
    
## calling method
file_path = "C://demo.xlsx" 
new_data = [
    {"Name": "Samyak", "Age": 12, "City": "yavatmal"},
    {"Name": "Chintu", "Age": 17, "City": "Goa"}]
    
insert_data_into_excel(file_path,new_data)